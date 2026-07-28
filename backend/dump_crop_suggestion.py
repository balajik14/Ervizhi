import openpyxl
import json
import os

wb = openpyxl.load_workbook('models_and_data/crop_suggestion.xlsx')
ws = wb.active

data = []
headers = [cell.value for cell in ws[1]]

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        item = {}
        for i, val in enumerate(row):
            item[headers[i]] = val
        data.append(item)

output_path = os.path.join('..', 'mobile', 'assets', 'crop_suggestion.json')
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"Dumped {len(data)} rows to {output_path}")
